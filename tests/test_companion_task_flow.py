"""Companion РОП «Задачи за день» — per-manager task flow across the day.

``day.team_task_flow`` reconstructs how many tasks each manager still had open at
each checkpoint (Bitrix keeps no such history), counts the day's new tasks and
sums talk time. These tests fake the Bitrix client and verify the filters it
issues, the created-before/closed-after reconstruction, the telephony exclusion,
and the truncation flag. The «встреч назначено» column comes from a different
source (the ТМ funnel's booking stage) and is assembled a level up, in
``service.get_team_task_flow`` — covered at the end.
"""

from __future__ import annotations

from collections.abc import AsyncIterator, Iterator
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Self
from zoneinfo import ZoneInfo

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from AtamuraOKK.bitrix import BitrixError
from AtamuraOKK.db.models.department import Department
from AtamuraOKK.db.models.manager import Manager
from AtamuraOKK.settings import settings
from AtamuraOKK.web.api.v1 import day, export, service
from AtamuraOKK.web.api.v1.okk import PeriodError
from AtamuraOKK.web.api.v1.schemas import TeamTaskFlow

pytestmark = pytest.mark.anyio

_TZ = ZoneInfo(settings.report_timezone)
_DAY = datetime(2026, 7, 2, tzinfo=_TZ)
_DAY_END = _DAY + timedelta(days=1)
_CHECKPOINTS = [_DAY.replace(hour=h) for h in (10, 14, 18)]


def _at(hour: int, minute: int = 0) -> str:
    return _DAY.replace(hour=hour, minute=minute).isoformat()


class FakeTaskBitrix:
    """Replays the three activity scans plus the telephony scan, by filter shape."""

    def __init__(
        self,
        still_open: list[dict[str, Any]] | None = None,
        closed: list[dict[str, Any]] | None = None,
        created: list[dict[str, Any]] | None = None,
        calls: list[dict[str, Any]] | None = None,
    ) -> None:
        self.still_open = still_open or []
        self.closed = closed or []
        self.created = created or []
        self.calls = calls or []
        self.filters: dict[str, dict[str, Any]] = {}
        self.scans: list[str] = []  # every scan issued, in order — one entry each

    def _rows_for(self, params: dict[str, Any]) -> list[dict[str, Any]]:
        flt = params.get("filter") or {}
        if flt.get("COMPLETED") == "N":
            self.filters["open"] = flt
            self.scans.append("open")
            return self.still_open
        if flt.get("COMPLETED") == "Y":
            self.filters["closed"] = flt
            self.scans.append("closed")
            return self.closed
        self.filters["created"] = flt
        self.scans.append("created")
        return self.created

    async def list(
        self,
        method: str,
        params: dict[str, Any] | None = None,
        *,
        max_items: int | None = None,
    ) -> AsyncIterator[dict[str, Any]]:
        """Yield the canned rows for whichever scan is being run."""
        params = params or {}
        if method == "voximplant.statistic.get":
            self.filters["calls"] = params.get("FILTER") or {}
            self.scans.append("calls")
            rows = self.calls
        else:
            assert method == "crm.activity.list"
            rows = self._rows_for(params)
        for i, row in enumerate(rows):
            if max_items is not None and i >= max_items:
                return
            yield row


def _task(
    uid: int,
    created: str,
    last_updated: str | None = None,
    deadline: str | None = None,
) -> dict[str, Any]:
    row = {"ID": "1", "RESPONSIBLE_ID": str(uid), "CREATED": created}
    if last_updated:
        row["LAST_UPDATED"] = last_updated
    if deadline:
        row["DEADLINE"] = deadline
    return row


def _call(
    uid: int,
    seconds: int,
    failed_code: str | None = None,
    at: str | None = None,
) -> dict[str, Any]:
    row = {
        "PORTAL_USER_ID": str(uid),
        "CALL_DURATION": str(seconds),
        "CALL_FAILED_CODE": failed_code or settings.ingest_success_code,
    }
    if at:
        row["CALL_START_DATE"] = at
    return row


async def _flow(bx: FakeTaskBitrix, uids: list[int]) -> tuple[dict[int, Any], bool]:
    return await day.team_task_flow(
        bx,  # type: ignore[arg-type]
        uids,
        _DAY,
        _DAY_END,
        _CHECKPOINTS,
        100,
    )


async def test_empty_roster_short_circuits() -> None:
    """No team members -> no Bitrix read at all."""
    bx = FakeTaskBitrix(still_open=[_task(5, _at(8))])
    counts, truncated = await _flow(bx, [])
    assert counts == {}
    assert truncated is False
    assert bx.filters == {}


async def test_still_open_task_counts_from_its_creation_on() -> None:
    """Created 12:00, never closed -> absent at 10:00, present at 14:00 and 18:00."""
    bx = FakeTaskBitrix(still_open=[_task(5, _at(12))])
    counts, _ = await _flow(bx, [5])
    assert counts[5].open_at == {10: 0, 14: 1, 18: 1}


async def test_closed_task_stops_counting_after_it_is_closed() -> None:
    """Created yesterday, closed 13:00 -> open at 10:00 only."""
    yesterday = (_DAY - timedelta(days=1)).replace(hour=9).isoformat()
    bx = FakeTaskBitrix(closed=[_task(5, yesterday, _at(13))])
    counts, _ = await _flow(bx, [5])
    assert counts[5].open_at == {10: 1, 14: 0, 18: 0}


async def test_task_closed_exactly_on_a_checkpoint_still_counts_there() -> None:
    """Closed at 14:00 sharp -> it was still open when the 14:00 срез was taken."""
    bx = FakeTaskBitrix(closed=[_task(5, _at(9), _at(14))])
    counts, _ = await _flow(bx, [5])
    assert counts[5].open_at == {10: 1, 14: 1, 18: 0}


async def test_backlog_declines_across_the_day() -> None:
    """The whole point of the table: остаток разгребается 3 -> 2 -> 1."""
    bx = FakeTaskBitrix(
        still_open=[_task(5, _at(8))],
        closed=[_task(5, _at(8), _at(11)), _task(5, _at(8), _at(15))],
    )
    counts, _ = await _flow(bx, [5])
    assert counts[5].open_at == {10: 3, 14: 2, 18: 1}


async def test_scans_exclude_telephony_and_scope_the_day() -> None:
    """Every scan drops the auto-logged call activities and stays in the day."""
    bx = FakeTaskBitrix(still_open=[_task(5, _at(8))])
    await _flow(bx, [5, 9])
    for key in ("open", "closed", "created"):
        assert bx.filters[key]["!PROVIDER_ID"] == day._TELEPHONY_PROVIDERS
        assert bx.filters[key]["RESPONSIBLE_ID"] == [5, 9]
    assert bx.filters["open"][">=DEADLINE"] == day._DEADLINE_FLOOR
    assert bx.filters["open"]["<DEADLINE"] == _DAY_END.isoformat()
    # only tasks touched at/after the first checkpoint can have been open at one
    assert bx.filters["closed"][">=LAST_UPDATED"] == _CHECKPOINTS[0].isoformat()
    assert bx.filters["created"][">=CREATED"] == _DAY.isoformat()
    assert bx.filters["created"]["<CREATED"] == _DAY_END.isoformat()


async def test_counts_new_tasks_and_talk_time_per_manager() -> None:
    """«Новых» and «время на линии» are attributed per manager, others zeroed."""
    bx = FakeTaskBitrix(
        created=[_task(5, _at(9)), _task(5, _at(16)), _task(9, _at(10))],
        calls=[_call(5, 120), _call(5, 300), _call(9, 60)],
    )
    counts, _ = await _flow(bx, [5, 9, 12])
    assert (counts[5].created, counts[5].talk_seconds) == (2, 420)
    assert (counts[9].created, counts[9].talk_seconds) == (1, 60)
    assert (counts[12].created, counts[12].talk_seconds) == (0, 0)
    assert counts[12].open_at == {10: 0, 14: 0, 18: 0}


async def test_unanswered_calls_do_not_count_as_time_on_the_line() -> None:
    """Only answered calls (success code) add to talk time."""
    bx = FakeTaskBitrix(calls=[_call(5, 120), _call(5, 45, failed_code="304")])
    counts, _ = await _flow(bx, [5])
    assert counts[5].talk_seconds == 120


async def test_rows_of_other_managers_are_ignored() -> None:
    """A row Bitrix returns outside the roster never lands on someone else."""
    bx = FakeTaskBitrix(still_open=[_task(7, _at(8))], created=[_task(7, _at(8))])
    counts, _ = await _flow(bx, [5])
    assert set(counts) == {5}
    assert counts[5].open_at == {10: 0, 14: 0, 18: 0}
    assert counts[5].created == 0


async def test_hitting_the_scan_cap_flags_truncation() -> None:
    """More rows than the cap -> counts under-report, truncated=True."""
    bx = FakeTaskBitrix(still_open=[_task(5, _at(8)) for _ in range(5)])
    counts, truncated = await day.team_task_flow(
        bx,  # type: ignore[arg-type]
        [5],
        _DAY,
        _DAY_END,
        _CHECKPOINTS,
        3,
    )
    assert truncated is True
    assert counts[5].open_at == {10: 3, 14: 3, 18: 3}


async def test_no_elapsed_checkpoints_short_circuits() -> None:
    """A day that hasn't reached its first срез yet reads nothing."""
    bx = FakeTaskBitrix(still_open=[_task(5, _at(8))])
    counts, truncated = await day.team_task_flow(
        bx,  # type: ignore[arg-type]
        [5],
        _DAY,
        _DAY_END,
        [],
        100,
    )
    assert counts == {}
    assert truncated is False
    assert bx.filters == {}


# --- «Встреч назначено» (service level) --------------------------------------
# The column comes from the ТМ funnel's booking stage, not from the activity
# scans, so it is assembled in ``service.get_team_task_flow`` rather than in
# ``day.team_task_flow``. These cover the three states the cabinet renders:
# a count, «—» off the ТМ department, and «—» when the stage read fails.


class FakeFlowBitrix(FakeTaskBitrix):
    """``FakeTaskBitrix`` plus the stage-history + deal reads «встреч назначено» makes.

    ``booked`` maps a deal id to the manager it is assigned to; those deals are
    replayed as this day's entrants into the booking stage, or set
    ``stage_rows`` to control the raw history (with its ``CREATED_TIME``) for the
    per-day split. ``stage_raises`` makes only that read fail, leaving the task
    scans intact; ``scan_raises`` kills the activity/telephony scans instead.
    """

    def __init__(
        self,
        *,
        booked: dict[int, int] | None = None,
        stage_raises: bool = False,
        scan_raises: bool = False,
        **kwargs: Any,
    ) -> None:
        super().__init__(**kwargs)
        self.booked = booked or {}
        self.stage_rows: list[dict[str, Any]] | None = None
        self.stage_raises = stage_raises
        self.scan_raises = scan_raises
        self.stage_filters: list[dict[str, Any]] = []

    async def __aenter__(self) -> Self:
        return self

    async def __aexit__(self, *exc: object) -> bool:
        return False

    async def call_raw(
        self,
        method: str,
        params: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """Replay the booking-stage entrants for the day, or fail on demand."""
        assert method == "crm.stagehistory.list"
        if self.stage_raises:
            raise BitrixError("ERR", "boom", method)
        self.stage_filters.append((params or {}).get("filter") or {})
        items = self.stage_rows
        if items is None:
            items = [{"OWNER_ID": str(d)} for d in self.booked]
        return {"result": {"items": items}}

    async def list(
        self,
        method: str,
        params: dict[str, Any] | None = None,
        *,
        max_items: int | None = None,
    ) -> AsyncIterator[dict[str, Any]]:
        """Resolve booked deals to their assignee; everything else is a task scan."""
        if method == "crm.deal.list":
            for did, uid in self.booked.items():
                yield {"ID": str(did), "ASSIGNED_BY_ID": str(uid)}
            return
        if self.scan_raises:
            raise BitrixError("ERR", "boom", method)
        async for row in super().list(method, params, max_items=max_items):
            yield row


async def _seed_team(
    session: AsyncSession,
    *,
    bitrix_id: int,
    uids: list[int],
) -> None:
    dept = Department(bitrix_id=bitrix_id, name=f"Отдел {bitrix_id}")
    session.add(dept)
    await session.flush()
    for uid in uids:
        session.add(
            Manager(bitrix_user_id=uid, name=f"М{uid}", department_id=dept.id),
        )
    await session.flush()


@pytest.fixture(autouse=True)
def _no_flow_caches() -> Iterator[None]:
    """Both views are TTL-cached; don't let one test read another's numbers."""
    service._task_flow_cache.clear()
    day._entrants_cache.clear()
    yield
    service._task_flow_cache.clear()
    day._entrants_cache.clear()


async def _team_flow(
    session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
    bx: FakeFlowBitrix,
    bitrix_id: int,
) -> TeamTaskFlow | None:
    monkeypatch.setattr(service, "BitrixClient", lambda *a, **k: bx)
    return await service.get_team_task_flow(session, bitrix_id, _DAY.date().isoformat())


async def test_meetings_set_counts_the_days_bookings_per_manager(
    dbsession: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Deals entering the booking stage that day land on their assignee."""
    dept = settings.companion_tm_department_id
    await _seed_team(dbsession, bitrix_id=dept, uids=[5, 7])
    bx = FakeFlowBitrix(booked={11: 5, 12: 5, 13: 7})
    flow = await _team_flow(dbsession, monkeypatch, bx, dept)

    assert flow is not None
    assert {r.manager.bitrix_user_id: r.meetings_set for r in flow.rows} == {5: 2, 7: 1}
    # Same stage and same day the «Важные цифры дня» tile counts.
    assert bx.stage_filters[0]["STAGE_ID"] == [settings.companion_meeting_set_stage_id]
    assert bx.stage_filters[0][">=CREATED_TIME"] == _DAY.date().isoformat()


async def test_manager_who_booked_nothing_reads_zero_not_null(
    dbsession: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Within ТМ a blank day is a real zero — null is reserved for 'no source'."""
    dept = settings.companion_tm_department_id
    await _seed_team(dbsession, bitrix_id=dept, uids=[5])
    flow = await _team_flow(dbsession, monkeypatch, FakeFlowBitrix(), dept)

    assert flow is not None
    assert flow.rows[0].meetings_set == 0


async def test_meeting_office_gets_null_not_a_misleading_zero(
    dbsession: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Outside ТМ the funnel doesn't apply: «—», and no Bitrix read is spent."""
    other = settings.companion_tm_department_id + 1
    await _seed_team(dbsession, bitrix_id=other, uids=[5])
    bx = FakeFlowBitrix(booked={11: 5})
    flow = await _team_flow(dbsession, monkeypatch, bx, other)

    assert flow is not None
    assert flow.rows[0].meetings_set is None
    assert bx.stage_filters == []


async def test_failed_stage_read_only_blanks_the_meetings_column(
    dbsession: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A dead stage-history read must not take the task columns down with it."""
    dept = settings.companion_tm_department_id
    await _seed_team(dbsession, bitrix_id=dept, uids=[5])
    bx = FakeFlowBitrix(
        stage_raises=True,
        still_open=[_task(5, _at(8))],
        created=[_task(5, _at(8))],
    )
    flow = await _team_flow(dbsession, monkeypatch, bx, dept)

    assert flow is not None
    assert flow.data_ready is True
    assert flow.rows[0].meetings_set is None
    assert flow.rows[0].created == 1
    assert [c.open_tasks for c in flow.rows[0].checkpoints] == [1, 1, 1]


# --- Range read (день / неделя / месяц одним набором сканов) ------------------
# The export needs the same per-day numbers over a whole period. Running the day
# view N times would be N× the Bitrix round-trips, so the range is read once and
# bucketed locally — these pin that it really is one read, and that a day inside
# the range still sees exactly what the day view would have shown it.

_PREV = _DAY - timedelta(days=1)


def _prev_at(hour: int) -> str:
    return _PREV.replace(hour=hour).isoformat()


def _spec(start: datetime, hours: tuple[int, ...] = (10, 14, 18)) -> day.TaskFlowDay:
    return day.TaskFlowDay(
        start,
        start + timedelta(days=1),
        [start.replace(hour=h) for h in hours],
    )


async def test_a_range_is_read_in_one_set_of_scans_not_one_per_day() -> None:
    """Two days must cost the same round-trips as one — that is the whole point."""
    bx = FakeTaskBitrix(still_open=[_task(5, _prev_at(8))])
    per_day, _ = await day.team_task_flow_range(
        bx,  # type: ignore[arg-type]
        [5],
        [_spec(_PREV), _spec(_DAY)],
        100,
    )
    assert sorted(per_day) == ["2026-07-01", "2026-07-02"]
    assert bx.scans == ["open", "closed", "created", "calls"]
    # The workload scan spans the range, and the closed floor is its first срез.
    assert bx.filters["open"]["<DEADLINE"] == (_DAY + timedelta(days=1)).isoformat()
    assert bx.filters["closed"][">=LAST_UPDATED"] == _prev_at(10)


async def test_a_day_only_counts_tasks_that_were_due_by_its_own_end() -> None:
    """The range scan is a superset; each day re-narrows it to its own workload."""
    bx = FakeTaskBitrix(
        still_open=[_task(5, _prev_at(8), deadline=_at(12))],  # due on day 2
    )
    per_day, _ = await day.team_task_flow_range(
        bx,  # type: ignore[arg-type]
        [5],
        [_spec(_PREV), _spec(_DAY)],
        100,
    )
    assert per_day["2026-07-01"][5].open_at == {10: 0, 14: 0, 18: 0}
    assert per_day["2026-07-02"][5].open_at == {10: 1, 14: 1, 18: 1}


async def test_new_tasks_and_talk_time_land_on_their_own_day() -> None:
    """Created and telephony rows are bucketed by their own timestamp."""
    bx = FakeTaskBitrix(
        created=[_task(5, _prev_at(9)), _task(5, _at(9)), _task(5, _at(11))],
        calls=[_call(5, 60, at=_prev_at(10)), _call(5, 180, at=_at(10))],
    )
    per_day, _ = await day.team_task_flow_range(
        bx,  # type: ignore[arg-type]
        [5],
        [_spec(_PREV), _spec(_DAY)],
        100,
    )
    assert per_day["2026-07-01"][5].created == 1
    assert per_day["2026-07-02"][5].created == 2
    assert per_day["2026-07-01"][5].talk_seconds == 60
    assert per_day["2026-07-02"][5].talk_seconds == 180


async def test_a_day_with_no_elapsed_checkpoint_is_left_out_of_the_range() -> None:
    """Today before 10:00 has nothing to report — no row, not a row of zeros."""
    bx = FakeTaskBitrix(still_open=[_task(5, _prev_at(8))])
    per_day, _ = await day.team_task_flow_range(
        bx,  # type: ignore[arg-type]
        [5],
        [_spec(_PREV), _spec(_DAY, hours=())],
        100,
    )
    assert sorted(per_day) == ["2026-07-01"]


async def test_bookings_split_per_day_and_dedupe_within_one() -> None:
    """A deal entering the stage twice in a day is one booking; on two days, two."""
    bx = FakeFlowBitrix()
    bx.stage_rows = [
        {"OWNER_ID": "11", "CREATED_TIME": _prev_at(9)},
        {"OWNER_ID": "11", "CREATED_TIME": _prev_at(15)},  # re-entry, same day
        {"OWNER_ID": "11", "CREATED_TIME": _at(9)},  # again the next day
        {"OWNER_ID": "12", "CREATED_TIME": _at(9)},
    ]
    bx.booked = {11: 5, 12: 7}
    per_day = await day.stage_entrants_by_assignee_per_day(
        bx,  # type: ignore[arg-type]
        settings.companion_meeting_set_stage_id,
        _PREV,
        _DAY + timedelta(days=1),
    )
    assert per_day["2026-07-01"] == {5: 1}
    assert per_day["2026-07-02"] == {5: 1, 7: 1}


# --- Excel-выгрузка «Задачи за день» ------------------------------------------


async def _task_flow_export(
    session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
    bx: FakeFlowBitrix,
    bitrix_id: int,
    period: str,
) -> Any:
    monkeypatch.setattr(export, "BitrixClient", lambda *a, **k: bx)
    return await export.get_task_flow_export(session, bitrix_id, period)


def _sheet(data: bytes, title: str) -> list[list[Any]]:
    wb = load_workbook(BytesIO(data))
    return [list(row) for row in wb[title].iter_rows(values_only=True)]


async def test_export_has_a_row_per_manager_per_day(
    dbsession: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A week export is the card's table repeated for every day of the period."""
    dept = settings.companion_tm_department_id
    await _seed_team(dbsession, bitrix_id=dept, uids=[5, 7])
    bx = FakeFlowBitrix(
        still_open=[_task(5, _prev_at(8))],
        created=[_task(5, _at(9))],
        calls=[_call(5, 120, at=_at(10))],
        booked={11: 5},
    )
    bx.stage_rows = [{"OWNER_ID": "11", "CREATED_TIME": _at(9)}]
    data = await _task_flow_export(
        dbsession,
        monkeypatch,
        bx,
        dept,
        "2026-07-01..2026-07-02",
    )

    assert data is not None
    assert data.data_ready is True
    assert data.with_meetings is True
    rows = _sheet(export.build_task_flow_workbook(data), "Задачи по дням")
    assert rows[0] == [
        "Дата",
        "Менеджер",
        "10:00",
        "14:00",
        "18:00",
        "Новых",
        "На линии",
        "Встреч назн.",
    ]
    assert len(rows) == 1 + 2 * 2  # header + 2 managers × 2 days
    # 2 июля у М5: остаток 1 на каждом срезе, 1 новая, 2 минуты, 1 встреча.
    july2 = [r for r in rows[1:] if r[0].date().isoformat() == "2026-07-02"]
    assert [r[1] for r in july2] == ["М5", "М7"]
    assert july2[0][2:] == [1, 1, 1, 1, "2:00", 1]


async def test_export_summary_averages_the_checkpoints_over_the_period(
    dbsession: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The rollup sheet is per manager, not per day — remainders average."""
    dept = settings.companion_tm_department_id
    await _seed_team(dbsession, bitrix_id=dept, uids=[5])
    # Открыта только со 2-го: 1-го числа остаток 0, 2-го — 1 → в среднем 0.5.
    bx = FakeFlowBitrix(still_open=[_task(5, _at(8))], created=[_task(5, _at(9))])
    data = await _task_flow_export(
        dbsession,
        monkeypatch,
        bx,
        dept,
        "2026-07-01..2026-07-02",
    )

    assert data is not None
    rows = _sheet(export.build_task_flow_workbook(data), "Сводка")
    assert rows[0][:2] == ["Менеджер", "Дней в периоде"]
    assert rows[1][:6] == ["М5", 2, 0.5, 0.5, 0.5, 1]
    assert ["Период", "2026-07-01..2026-07-02"] in [r[:2] for r in rows]


async def test_meeting_office_export_has_no_meetings_column(
    dbsession: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Same rule as the card: no ТМ funnel, no «Встреч назн.» column at all."""
    other = settings.companion_tm_department_id + 1
    await _seed_team(dbsession, bitrix_id=other, uids=[5])
    bx = FakeFlowBitrix(still_open=[_task(5, _at(8))])
    data = await _task_flow_export(dbsession, monkeypatch, bx, other, "2026-07-02")

    assert data is not None
    assert data.with_meetings is False
    assert bx.stage_filters == []
    rows = _sheet(export.build_task_flow_workbook(data), "Задачи по дням")
    assert "Встреч назн." not in rows[0]


async def test_export_refuses_a_period_longer_than_the_cap(
    dbsession: AsyncSession,
) -> None:
    """A year-long range would run for many minutes — refuse it as a bad period."""
    dept = settings.companion_tm_department_id
    await _seed_team(dbsession, bitrix_id=dept, uids=[5])
    with pytest.raises(PeriodError, match="at most"):
        await export.get_task_flow_export(dbsession, dept, "2026-01-01..2026-12-31")


async def test_unreadable_bitrix_marks_the_export_not_ready(
    dbsession: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A dead Bitrix must be said out loud in the file, not shipped as empty rows."""
    dept = settings.companion_tm_department_id
    await _seed_team(dbsession, bitrix_id=dept, uids=[5])
    bx = FakeFlowBitrix(scan_raises=True)
    data = await _task_flow_export(dbsession, monkeypatch, bx, dept, "2026-07-02")

    assert data is not None
    assert data.data_ready is False
    assert data.rows == []
    rows = _sheet(export.build_task_flow_workbook(data), "Сводка")
    assert any(r[0] == "⚠ Bitrix не ответил" for r in rows)
