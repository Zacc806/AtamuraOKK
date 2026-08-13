"""РОП Excel export: one row per scored item of a department, for a period.

The Команда screen answers "how is the team doing"; a РОП also needs the raw
list behind that number — who, when, what score — to hand around outside the
cabinet. This module builds that .xlsx.

It reads the same sources as the team summary (``call_scores_latest`` for
calls, ``meetings`` for meeting offices) and applies the same appeal
overrides, so an exported percent equals the one the cabinet shows. A
department exports whatever it produces: ТМ → calls, ОП-офис → meetings.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.sql import text

from AtamuraOKK.bitrix import BitrixClient, BitrixError, crm_card_url
from AtamuraOKK.db.models.department import Department
from AtamuraOKK.db.models.manager import Manager
from AtamuraOKK.db.models.meeting import Meeting
from AtamuraOKK.settings import settings
from AtamuraOKK.web.api.v1 import day, okk
from AtamuraOKK.web.api.v1.service import (
    _apply_score_overrides,
    _counts_in_score,
    _full_name,
    _meetings_score_from,
    _okk_from_rows,
    _roster_names,
)

CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_HEADER_FILL = PatternFill("solid", fgColor="0E2A40")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


@dataclass
class ExportRow:
    """One scored item, already localized to the report timezone."""

    manager: str
    at: datetime | None
    percent: float | None
    extra: list[Any] = field(default_factory=list)


@dataclass
class TeamExport:
    """Everything the workbook needs: identity, the detail rows, the rollup."""

    department_name: str
    department_bitrix_id: int
    period: str
    kind: str  # "calls" | "meetings"
    rows: list[ExportRow]
    summary: list[list[Any]]


def _local(value: datetime | None, tz: ZoneInfo) -> datetime | None:
    """Report-timezone wall clock, tz-naive — Excel has no timezone concept."""
    if value is None:
        return None
    return value.astimezone(tz).replace(tzinfo=None)


def _duration(seconds: int | None) -> str | None:
    if seconds is None:
        return None
    return f"{seconds // 60}:{seconds % 60:02d}"


_CALL_COLUMNS = (
    ("Менеджер", 28),
    ("Дата и время", 18),
    ("Длительность", 13),
    ("Тип звонка", 20),
    ("Оценка, %", 11),
    ("ОКК (1–5)", 10),
    ("Зона", 13),
    ("Учитывается в ОКК", 18),
    ("Карточка CRM", 42),
)

_MEETING_COLUMNS = (
    ("Менеджер", 28),
    ("Дата и время", 18),
    ("Длительность", 13),
    ("Запись", 40),
    ("Оценка, %", 11),
    ("Результат", 14),
    ("На проверке", 13),
)

_CALL_SUMMARY_COLUMNS = (
    ("Менеджер", 28),
    ("Звонков в зачёт", 16),
    ("Средний балл, %", 16),
    ("ОКК (1–5)", 10),
    ("Зона", 13),
)

_MEETING_SUMMARY_COLUMNS = (
    ("Менеджер", 28),
    ("Встреч оценено", 16),
    ("Средний балл, %", 16),
    ("Сдано", 10),
    ("Не сдано", 10),
)

_ZONE_RU = {
    "strong": "сильная",
    "normal": "норма",
    "borderline": "погранич.",
    "risk": "риск",
}


async def _call_rows(
    session: AsyncSession,
    department_id: int,
    start: datetime,
    end: datetime,
    tz: ZoneInfo,
) -> tuple[list[ExportRow], list[list[Any]]]:
    """Detail rows + per-manager rollup for a calling department."""
    rows = (
        await session.execute(
            text(
                "SELECT call_id, manager_bitrix_user_id, manager_name, started_at, "
                "duration_sec, percent, zone, call_type, is_qualification_call, "
                "crm_entity_type, crm_entity_id FROM call_scores_latest "
                "WHERE department_id = :dept "
                "AND started_at >= :start AND started_at < :end "
                "ORDER BY manager_name NULLS LAST, started_at",
            ),
            {"dept": department_id, "start": start, "end": end},
        )
    ).all()
    rows = await _apply_score_overrides(session, rows)

    detail: list[ExportRow] = []
    for r in rows:
        percent = float(r.percent) if r.percent is not None else None
        counts = _counts_in_score(r)
        detail.append(
            ExportRow(
                manager=r.manager_name or f"ID {r.manager_bitrix_user_id}",
                at=_local(r.started_at, tz),
                percent=percent,
                extra=[
                    _duration(r.duration_sec),
                    r.call_type,
                    okk.okk_5(percent) if counts else None,
                    _ZONE_RU.get(r.zone or "", r.zone),
                    "да" if counts else "нет",
                    crm_card_url(r.crm_entity_type, r.crm_entity_id),
                ],
            ),
        )

    by_manager: dict[str, list[Any]] = {}
    for r in rows:
        name = r.manager_name or f"ID {r.manager_bitrix_user_id}"
        by_manager.setdefault(name, []).append(r)

    summary: list[list[Any]] = []
    for name, mgr_rows in by_manager.items():
        score, _zones, n = _okk_from_rows(mgr_rows)
        zone = _ZONE_RU.get(score.zone or "", "—")
        summary.append([name, n, score.percent, score.score_5, zone])
    summary.sort(key=lambda row: (row[2] is None, -(row[2] or 0.0)))
    return detail, summary


async def _meeting_rows(
    session: AsyncSession,
    department_id: int,
    start: datetime,
    end: datetime,
    tz: ZoneInfo,
) -> tuple[list[ExportRow], list[list[Any]]]:
    """Detail rows + per-manager rollup for a meeting office."""
    pairs = (
        await session.execute(
            select(Meeting, Manager)
            .join(Manager, Meeting.manager_id == Manager.id)
            .where(
                Manager.department_id == department_id,
                Meeting.meeting_at >= start,
                Meeting.meeting_at < end,
            )
            .order_by(Meeting.meeting_at),
        )
    ).all()

    detail: list[ExportRow] = []
    by_manager: dict[str, list[Meeting]] = {}
    for meeting, mgr in pairs:
        name = _full_name(mgr) or f"ID {mgr.bitrix_user_id}"
        by_manager.setdefault(name, []).append(meeting)
        detail.append(
            ExportRow(
                manager=name,
                at=_local(meeting.meeting_at, tz),
                percent=(
                    float(meeting.score_pct) if meeting.score_pct is not None else None
                ),
                extra=[
                    _duration(meeting.duration_sec),
                    meeting.name,
                    (
                        "—"
                        if meeting.passed is None
                        else ("сдано" if meeting.passed else "не сдано")
                    ),
                    "да" if meeting.needs_human_review else "",
                ],
            ),
        )
    summary: list[list[Any]] = []
    for name, meetings in by_manager.items():
        agg = _meetings_score_from(meetings)
        summary.append(
            [name, agg.meetings_scored, agg.avg_score_pct, agg.passed, agg.failed],
        )
    summary.sort(key=lambda row: (row[2] is None, -(row[2] or 0.0)))
    return detail, summary


async def get_team_export(
    session: AsyncSession,
    department_bitrix_id: int,
    period: str | None,
) -> TeamExport | None:
    """Collect the department's scored items for a period, or None if unknown.

    ``period`` accepts everything ``okk.parse_period`` does — a month
    (``YYYY-MM``), a week (``YYYY-MM-DD..YYYY-MM-DD``) or a single day.
    """
    department = await session.scalar(
        select(Department).where(Department.bitrix_id == department_bitrix_id),
    )
    if department is None:
        return None

    start, end, label = okk.parse_period(period)
    tz = ZoneInfo(settings.report_timezone)
    is_tm = department.bitrix_id == settings.companion_tm_department_id
    if is_tm:
        detail, summary = await _call_rows(session, department.id, start, end, tz)
    else:
        detail, summary = await _meeting_rows(session, department.id, start, end, tz)
    return TeamExport(
        department_name=department.name or f"ID {department_bitrix_id}",
        department_bitrix_id=department_bitrix_id,
        period=label,
        kind="calls" if is_tm else "meetings",
        rows=detail,
        summary=summary,
    )


def _write_header(ws: Worksheet, columns: tuple[tuple[str, int], ...]) -> None:
    ws.append([title for title, _ in columns])
    for idx, (_title, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def _format_column(ws: Worksheet, column: int, fmt: str) -> None:
    for excel_row in ws.iter_rows(min_row=2, min_col=column, max_col=column):
        excel_row[0].number_format = fmt


def build_workbook(export: TeamExport) -> bytes:
    """Render the export as an .xlsx byte string."""
    wb = Workbook()
    ws = wb.active
    meetings = export.kind == "meetings"
    ws.title = "Встречи" if meetings else "Звонки"
    columns = _MEETING_COLUMNS if meetings else _CALL_COLUMNS
    _write_header(ws, columns)

    for row in export.rows:
        if meetings:
            duration, name, verdict, review = row.extra
            ws.append(
                [row.manager, row.at, duration, name, row.percent, verdict, review],
            )
        else:
            duration, call_type, okk_5, zone, counts, url = row.extra
            ws.append(
                [
                    row.manager,
                    row.at,
                    duration,
                    call_type,
                    row.percent,
                    okk_5,
                    zone,
                    counts,
                    url,
                ],
            )
    _format_column(ws, 2, "DD.MM.YYYY HH:MM")
    _format_column(ws, 5, "0.0")  # оценка, % — как в кабинете, один знак
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"

    summary_ws = wb.create_sheet("Сводка")
    summary_columns = _MEETING_SUMMARY_COLUMNS if meetings else _CALL_SUMMARY_COLUMNS
    _write_header(summary_ws, summary_columns)
    for summary_row in export.summary:
        summary_ws.append(summary_row)
    _format_column(summary_ws, 3, "0.0")
    summary_ws.append([])
    summary_ws.append(["Отдел", export.department_name])
    summary_ws.append(["Период", export.period])
    return _to_bytes(wb)


def _to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def filename_for(export: TeamExport) -> str:
    """ASCII filename — Cyrillic department names break naive HTTP clients."""
    period = export.period.replace("..", "_")
    kind = "meetings" if export.kind == "meetings" else "calls"
    return f"okk-{kind}-dept{export.department_bitrix_id}-{period}.xlsx"


# --- «Задачи за день» export --------------------------------------------------
# The same table the Команда card shows, for a day / week / month. Unlike the
# card (which is one day, live), the range is read in a single set of scans and
# split per day — see ``day.team_task_flow_range`` for why that matters.


@dataclass
class TaskFlowExportRow:
    """One manager on one day: остаток по срезам, новые, на линии, встречи."""

    date: datetime
    manager: str
    open_at: dict[int, int | None]
    created: int
    talk_seconds: int
    meetings_set: int | None


@dataclass
class TaskFlowExport:
    """Everything the «Задачи за день» workbook needs."""

    department_name: str
    department_bitrix_id: int
    period: str
    checkpoint_hours: list[int]
    rows: list[TaskFlowExportRow]
    with_meetings: bool
    truncated: bool
    data_ready: bool


def _export_days(start: datetime, end: datetime, tz: ZoneInfo) -> list[day.TaskFlowDay]:
    """Split a period into days, each carrying only the срезы that have elapsed.

    A future day, or today before its first срез, yields no checkpoints and is
    dropped by the range read — the export then simply has no row for it, which
    is honest: nothing has been measured yet.
    """
    now = datetime.now(tz=tz)
    hours = sorted(settings.companion_task_checkpoints)
    days: list[day.TaskFlowDay] = []
    cursor = start
    while cursor < end:
        nxt = cursor + timedelta(days=1)
        checkpoints = [
            cp for cp in (cursor.replace(hour=h) for h in hours) if cp <= now
        ]
        if checkpoints:
            days.append(day.TaskFlowDay(cursor, nxt, checkpoints))
        cursor = nxt
    return days


async def get_task_flow_export(
    session: AsyncSession,
    department_bitrix_id: int,
    period: str | None,
) -> TaskFlowExport | None:
    """Collect «Задачи за день» for every day of a period, or None if unknown.

    ``period`` accepts everything ``okk.parse_period`` does — a single day, a
    week (``YYYY-MM-DD..YYYY-MM-DD``) or a month. Longer than
    ``companion_task_flow_export_max_days`` is refused as a period error rather
    than left to run for minutes. ``data_ready=False`` means Bitrix was
    unreadable and the caller should say so instead of shipping an empty sheet.
    """
    department = await session.scalar(
        select(Department).where(Department.bitrix_id == department_bitrix_id),
    )
    if department is None:
        return None

    start, end, label = okk.parse_period(period)
    span = (end - start).days
    if span > settings.companion_task_flow_export_max_days:
        raise okk.PeriodError(
            f"period covers {span} days; the task-flow export allows at most "
            f"{settings.companion_task_flow_export_max_days}",
        )

    tz = ZoneInfo(settings.report_timezone)
    days = _export_days(start, end, tz)
    names = await _roster_names(session, department.id)
    with_meetings = department_bitrix_id == settings.companion_tm_department_id

    per_day: dict[str, dict[int, day.TaskFlowCounts]] = {}
    meetings: dict[str, dict[int, int]] = {}
    truncated = False
    data_ready = True
    if days and names:
        try:
            async with BitrixClient() as bx:
                per_day, truncated = await day.team_task_flow_range(
                    bx,
                    sorted(names),
                    days,
                    settings.companion_task_flow_export_max_scan,
                )
                if with_meetings:
                    try:
                        meetings = await day.stage_entrants_by_assignee_per_day(
                            bx,
                            settings.companion_meeting_set_stage_id,
                            start,
                            end,
                        )
                    except BitrixError as exc:
                        logger.warning(
                            "Task-flow export meetings read failed for dept {d}: {e}",
                            d=department_bitrix_id,
                            e=exc,
                        )
                        with_meetings = False
        except BitrixError as exc:
            data_ready = False
            logger.warning(
                "Task-flow export Bitrix read failed for dept {d}: {e}",
                d=department_bitrix_id,
                e=exc,
            )

    hours = sorted(settings.companion_task_checkpoints)
    elapsed = {
        d.start.date().isoformat(): {cp.hour for cp in d.checkpoints} for d in days
    }
    rows: list[TaskFlowExportRow] = []
    for label_day in sorted(per_day):
        counts = per_day[label_day]
        seen = elapsed.get(label_day, set())
        for uid in sorted(counts, key=lambda u: (names.get(u) or "").lower()):
            row = counts[uid]
            rows.append(
                TaskFlowExportRow(
                    date=datetime.fromisoformat(label_day),
                    manager=names.get(uid) or f"ID {uid}",
                    open_at={
                        h: (row.open_at.get(h) if h in seen else None) for h in hours
                    },
                    created=row.created,
                    talk_seconds=row.talk_seconds,
                    meetings_set=(
                        meetings.get(label_day, {}).get(uid, 0)
                        if with_meetings
                        else None
                    ),
                ),
            )

    return TaskFlowExport(
        department_name=department.name or f"ID {department_bitrix_id}",
        department_bitrix_id=department_bitrix_id,
        period=label,
        checkpoint_hours=hours,
        rows=rows,
        with_meetings=with_meetings,
        truncated=truncated,
        data_ready=data_ready,
    )


def _task_flow_columns(export: TaskFlowExport) -> tuple[tuple[str, int], ...]:
    columns: list[tuple[str, int]] = [("Дата", 12), ("Менеджер", 28)]
    columns += [(f"{h:02d}:00", 9) for h in export.checkpoint_hours]
    columns += [("Новых", 9), ("На линии", 11)]
    if export.with_meetings:
        columns.append(("Встреч назн.", 13))
    return tuple(columns)


def _task_flow_summary(export: TaskFlowExport) -> list[list[Any]]:
    """Per-manager rollup: средний остаток по срезам, суммы за период.

    The checkpoint columns average over the days that actually have a number, so
    a period containing today (whose 18:00 hasn't arrived) doesn't drag the mean
    down with a missing срез counted as zero.
    """
    by_manager: dict[str, list[TaskFlowExportRow]] = {}
    for row in export.rows:
        by_manager.setdefault(row.manager, []).append(row)

    summary: list[list[Any]] = []
    for name, rows in by_manager.items():
        line: list[Any] = [name, len(rows)]
        for hour in export.checkpoint_hours:
            seen = [n for n in (r.open_at.get(hour) for r in rows) if n is not None]
            line.append(round(sum(seen) / len(seen), 1) if seen else None)
        line.append(sum(r.created for r in rows))
        line.append(_duration(sum(r.talk_seconds for r in rows)))
        if export.with_meetings:
            line.append(sum(r.meetings_set or 0 for r in rows))
        summary.append(line)
    # Кто хуже разгребает — сверху: по среднему остатку на последнем срезе.
    last = len(export.checkpoint_hours) + 1
    summary.sort(key=lambda line: -(line[last] or 0))
    return summary


def _task_flow_summary_columns(export: TaskFlowExport) -> tuple[tuple[str, int], ...]:
    columns: list[tuple[str, int]] = [("Менеджер", 28), ("Дней в периоде", 14)]
    columns += [(f"Ср. остаток {h:02d}:00", 16) for h in export.checkpoint_hours]
    columns += [("Новых задач", 13), ("На линии", 11)]
    if export.with_meetings:
        columns.append(("Встреч назначено", 17))
    return tuple(columns)


def build_task_flow_workbook(export: TaskFlowExport) -> bytes:
    """Render «Задачи за день» for the period as an .xlsx byte string."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Задачи по дням"
    columns = _task_flow_columns(export)
    _write_header(ws, columns)

    for row in export.rows:
        line: list[Any] = [row.date, row.manager]
        line += [row.open_at.get(h) for h in export.checkpoint_hours]
        line += [row.created, _duration(row.talk_seconds)]
        if export.with_meetings:
            line.append(row.meetings_set)
        ws.append(line)
    _format_column(ws, 1, "DD.MM.YYYY")
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"

    summary_ws = wb.create_sheet("Сводка")
    _write_header(summary_ws, _task_flow_summary_columns(export))
    for summary_row in _task_flow_summary(export):
        summary_ws.append(summary_row)
    summary_ws.append([])
    summary_ws.append(["Отдел", export.department_name])
    summary_ws.append(["Период", export.period])
    if not export.data_ready:
        summary_ws.append(["⚠ Bitrix не ответил", "числа за период не прочитаны"])
    if export.truncated:
        summary_ws.append(["⚠ Слишком много дел", "часть не попала в подсчёт"])
    return _to_bytes(wb)


def task_flow_filename_for(export: TaskFlowExport) -> str:
    """ASCII filename — Cyrillic department names break naive HTTP clients."""
    period = export.period.replace("..", "_")
    return f"okk-taskflow-dept{export.department_bitrix_id}-{period}.xlsx"
