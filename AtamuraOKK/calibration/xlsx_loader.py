"""Load human OKK scores from the meeting checklist xlsx (calibration ground truth).

Verified layout of ``Чек лист встречи ОП - Январь.xlsx`` (per manager sheet,
sheet "Сводная" is a summary and is skipped):

- Row 1/2/3/4: per-call reviewer / date / duration / CRM-deal URL.
- Row 5: column headers.
- Rows 6-26: the 20 criteria. Column B = criterion number (row 14 is an
  unnumbered sub-note and is skipped); the per-call score sits one column right
  of the call's base column.
- Row 27: per-call raw total (0-50).
- Each call occupies a 3-column group (да/нет, оценка, комментарий) starting at
  column E (index 5): base columns 5, 8, 11, ...

Requires the ``calib`` dependency group (openpyxl).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path

_SUMMARY_SHEET = "Сводная"
_FIRST_CALL_COL = 5  # column E
_GROUP_STRIDE = 3
_CRM_ROW = 4
_TOTAL_ROW = 27
_REVIEWER_ROW = 1
_DATE_ROW = 2
_DURATION_ROW = 3
_FIRST_CRIT_ROW = 6
_LAST_CRIT_ROW = 26
_NUM_COL = 2  # column B (criterion number)
_DEAL_RE = re.compile(r"/deal/details/(\d+)")


@dataclass(slots=True)
class HumanCall:
    """One human-scored meeting from the xlsx."""

    manager: str
    reviewer: str | None
    crm_deal_id: int | None
    crm_url: str | None
    raw_total: int | None
    per_criterion: dict[int, int] = field(default_factory=dict)


def _deal_id(url: object) -> int | None:
    if not isinstance(url, str):
        return None
    match = _DEAL_RE.search(url)
    return int(match.group(1)) if match else None


def _as_int(value: object) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    return None


def _as_str(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def load_human_calls(path: str | Path) -> list[HumanCall]:
    """Parse every human-scored meeting from the checklist workbook.

    :param path: path to the OKK meeting checklist .xlsx.
    :returns: one :class:`HumanCall` per scored call across all manager sheets.
    """
    from openpyxl import load_workbook  # noqa: PLC0415

    workbook = load_workbook(path, data_only=True)
    calls: list[HumanCall] = []
    for sheet in workbook.worksheets:
        if sheet.title == _SUMMARY_SHEET:
            continue
        calls.extend(_parse_sheet(sheet))
    workbook.close()
    return calls


def _crit_rows(sheet: object) -> list[tuple[int, int]]:
    """Return (row, criterion_number) for each numbered criterion row."""
    rows: list[tuple[int, int]] = []
    for row in range(_FIRST_CRIT_ROW, _LAST_CRIT_ROW + 1):
        number = _as_int(sheet.cell(row=row, column=_NUM_COL).value)  # type: ignore[attr-defined]
        if number is not None and 1 <= number <= 20:
            rows.append((row, number))
    return rows


def _parse_sheet(sheet: object) -> list[HumanCall]:
    manager = str(sheet.title)  # type: ignore[attr-defined]
    max_col: int = sheet.max_column  # type: ignore[attr-defined]
    crit_rows = _crit_rows(sheet)
    calls: list[HumanCall] = []

    base = _FIRST_CALL_COL
    while base <= max_col:
        crm_url = sheet.cell(row=_CRM_ROW, column=base).value  # type: ignore[attr-defined]
        total = sheet.cell(row=_TOTAL_ROW, column=base).value  # type: ignore[attr-defined]
        if crm_url is None and total is None:
            base += _GROUP_STRIDE
            continue

        score_col = base + 1
        per_criterion: dict[int, int] = {}
        for row, number in crit_rows:
            score = _as_int(sheet.cell(row=row, column=score_col).value)  # type: ignore[attr-defined]
            if score is not None:
                per_criterion[number] = score

        calls.append(
            HumanCall(
                manager=manager,
                reviewer=_as_str(sheet.cell(row=_REVIEWER_ROW, column=base).value),  # type: ignore[attr-defined]
                crm_deal_id=_deal_id(crm_url),
                crm_url=_as_str(crm_url),
                raw_total=_as_int(total),
                per_criterion=per_criterion,
            ),
        )
        base += _GROUP_STRIDE

    return calls
